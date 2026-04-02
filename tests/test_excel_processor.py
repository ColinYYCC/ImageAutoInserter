"""
Excel 处理器单元测试
"""

import pytest
from pathlib import Path
from openpyxl import Workbook
from src.core.excel_processor import ExcelProcessor, SheetInfo


class TestExcelProcessor:
    """Excel 处理器测试类"""
    
    @pytest.fixture
    def sample_excel(self, tmp_path):
        """创建示例 Excel 文件"""
        # 创建测试 Excel 文件
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # 添加表头（在第 1 行）
        ws.cell(row=1, column=1, value="序号")
        ws.cell(row=1, column=2, value="商品编码")
        ws.cell(row=1, column=3, value="名称")
        ws.cell(row=1, column=4, value="价格")
        
        # 添加数据（第 2-4 行）
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value="C001")
        ws.cell(row=2, column=3, value="产品 A")
        ws.cell(row=2, column=4, value=100)
        
        ws.cell(row=3, column=1, value=2)
        ws.cell(row=3, column=2, value="C002")
        ws.cell(row=3, column=3, value="产品 B")
        ws.cell(row=3, column=4, value=200)
        
        ws.cell(row=4, column=1, value=3)
        ws.cell(row=4, column=2, value="C003")
        ws.cell(row=4, column=3, value="产品 C")
        ws.cell(row=4, column=4, value=300)
        
        # 保存文件
        excel_path = tmp_path / "test.xlsx"
        wb.save(str(excel_path))
        
        return excel_path
    
    @pytest.fixture
    def excel_with_header_offset(self, tmp_path):
        """创建表头不在第一行的 Excel 文件"""
        wb = Workbook()
        ws = wb.active
        
        # 第 1 行：标题
        ws.cell(row=1, column=1, value="销售报表")
        
        # 第 2 行：空行
        # 第 3 行：表头
        ws.cell(row=3, column=1, value="序号")
        ws.cell(row=3, column=2, value="商品编码")
        ws.cell(row=3, column=3, value="名称")
        
        # 第 4-5 行：数据
        ws.cell(row=4, column=1, value=1)
        ws.cell(row=4, column=2, value="C001")
        ws.cell(row=4, column=3, value="产品 A")
        
        ws.cell(row=5, column=1, value=2)
        ws.cell(row=5, column=2, value="C002")
        ws.cell(row=5, column=3, value="产品 B")
        
        excel_path = tmp_path / "test_offset.xlsx"
        wb.save(str(excel_path))
        
        return excel_path
    
    def test_init_file_not_exists(self):
        """测试初始化时文件不存在"""
        with pytest.raises(FileNotFoundError):
            ExcelProcessor("/non/existent/file.xlsx")
    
    def test_init_invalid_format(self, tmp_path):
        """测试初始化时无效的文件格式"""
        # 创建 .xls 文件（不支持的格式）
        xls_path = tmp_path / "test.xls"
        xls_path.touch()
        
        with pytest.raises(ValueError, match="不支持的文件格式"):
            ExcelProcessor(str(xls_path))
    
    def test_init_success(self, sample_excel):
        """测试初始化成功"""
        processor = ExcelProcessor(str(sample_excel))
        assert processor is not None
        assert processor.file_path == sample_excel
    
    def test_find_sheet_with_product_code(self, sample_excel):
        """测试查找包含商品编码的工作表"""
        processor = ExcelProcessor(str(sample_excel))
        
        sheet_info = processor.find_sheet_with_product_code()
        
        assert sheet_info is not None
        assert sheet_info.name == "Sheet1"
        assert sheet_info.header_row == 1
        assert sheet_info.product_code_column == 2
        assert len(sheet_info.data_rows) == 3
    
    def test_find_sheet_with_offset_header(self, excel_with_header_offset):
        """测试查找表头不在第一行的工作表"""
        processor = ExcelProcessor(str(excel_with_header_offset))
        
        sheet_info = processor.find_sheet_with_product_code()
        
        assert sheet_info is not None
        assert sheet_info.header_row == 3  # 表头在第 3 行
        assert sheet_info.product_code_column == 2
        assert len(sheet_info.data_rows) == 2  # 第 4-5 行
    
    def test_get_product_code(self, sample_excel):
        """测试获取商品编码"""
        processor = ExcelProcessor(str(sample_excel))
        processor.find_sheet_with_product_code()
        
        # 获取第 2 行的商品编码
        code = processor.get_product_code(2)
        assert code == "C001"
        
        # 获取第 3 行的商品编码
        code = processor.get_product_code(3)
        assert code == "C002"
    
    def test_get_product_codes(self, sample_excel):
        """测试获取所有商品编码"""
        processor = ExcelProcessor(str(sample_excel))
        processor.find_sheet_with_product_code()
        
        codes = processor.get_product_codes()
        
        assert len(codes) == 3
        assert "C001" in codes
        assert "C002" in codes
        assert "C003" in codes
    
    def test_add_picture_columns(self, sample_excel):
        """测试添加 Picture 列"""
        # 使用 read_only=False 以允许写入
        processor = ExcelProcessor(str(sample_excel), read_only=False)
        processor.find_sheet_with_product_code()
        
        # 添加 Picture 列
        result = processor.add_picture_columns()
        
        # 验证返回类型
        from src.core.picture_variant import ColumnAdditionResult
        assert isinstance(result, ColumnAdditionResult)
        
        # 验证添加了 3 列
        assert result.total_columns == 3
        assert len(result.added_columns) == 3
        assert len(result.existing_columns) == 0
        
        # 验证列号
        added_dict = {col.original_header: col.column_index for col in result.added_columns}
        assert "Picture 1" in added_dict
        assert "Picture 2" in added_dict
        assert "Picture 3" in added_dict
        
        # 验证表头已设置
        worksheet = processor.workbook["Sheet1"]
        assert worksheet.cell(row=1, column=added_dict["Picture 1"]).value == "Picture 1"
        assert worksheet.cell(row=1, column=added_dict["Picture 2"]).value == "Picture 2"
        assert worksheet.cell(row=1, column=added_dict["Picture 3"]).value == "Picture 3"
    
    def test_save(self, sample_excel, tmp_path):
        """测试保存文件"""
        # 使用 read_only=False 以允许写入
        processor = ExcelProcessor(str(sample_excel), read_only=False)
        processor.find_sheet_with_product_code()
        
        # 保存到新路径
        output_path = tmp_path / "output.xlsx"
        saved_path = processor.save(str(output_path))
        
        assert saved_path == output_path
        assert saved_path.exists()
    
    def test_save_auto_path(self, sample_excel):
        """测试自动保存路径"""
        # 使用 read_only=False 以允许写入
        processor = ExcelProcessor(str(sample_excel), read_only=False)
        processor.find_sheet_with_product_code()
        
        # 自动保存（不指定路径）
        saved_path = processor.save()
        
        assert saved_path.name == "test_含图.xlsx"
        assert saved_path.exists()
    
    def test_context_manager(self, sample_excel):
        """测试上下文管理器"""
        with ExcelProcessor(str(sample_excel)) as processor:
            sheet_info = processor.find_sheet_with_product_code()
            assert sheet_info is not None
        
        # 上下文管理器应自动关闭
    
    def test_highlight_empty_product_codes(self, sample_excel):
        """测试高亮未匹配图片的商品编码"""
        # 使用 read_only=False 以允许写入
        processor = ExcelProcessor(str(sample_excel), read_only=False)
        processor.find_sheet_with_product_code()
        
        # 高亮第 2 行和第 3 行
        processor.highlight_empty_product_codes([2, 3])
        
        # 验证单元格样式
        worksheet = processor.workbook["Sheet1"]
        product_code_column = processor.sheet_info.product_code_column
        
        # 检查第 2 行
        cell_row2 = worksheet.cell(row=2, column=product_code_column)
        assert cell_row2.fill.start_color.rgb == "00FFFF00"  # 黄色
        assert cell_row2.fill.fill_type == "solid"
        
        # 检查第 3 行
        cell_row3 = worksheet.cell(row=3, column=product_code_column)
        assert cell_row3.fill.start_color.rgb == "00FFFF00"  # 黄色
        assert cell_row3.fill.fill_type == "solid"
        
        # 检查第 4 行（未高亮）
        cell_row4 = worksheet.cell(row=4, column=product_code_column)
        assert cell_row4.fill.start_color.rgb == "00000000"  # 无填充
    
    def test_highlight_empty_product_codes_custom_color(self, sample_excel):
        """测试自定义颜色高亮"""
        processor = ExcelProcessor(str(sample_excel), read_only=False)
        processor.find_sheet_with_product_code()
        
        # 使用红色高亮
        processor.highlight_empty_product_codes([2], fill_color="FF0000")
        
        worksheet = processor.workbook["Sheet1"]
        product_code_column = processor.sheet_info.product_code_column
        cell = worksheet.cell(row=2, column=product_code_column)
        
        assert cell.fill.start_color.rgb == "00FF0000"  # 红色
