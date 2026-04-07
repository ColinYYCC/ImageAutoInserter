"""
流程编排器单元测试
测试覆盖：流程协调、进度报告、错误处理等
"""

import pytest
import tempfile
import threading
from pathlib import Path
from unittest.mock import Mock, MagicMock, patch, call
from datetime import datetime

from src.core.pipeline.orchestrator import (
    ProcessOrchestrator,
    ProcessContext,
    ThreadSafeProgressReporter
)
from src.core.models import ProcessResult, ProgressInfo, ErrorRecord


class TestThreadSafeProgressReporter:
    """线程安全进度报告器测试"""

    def test_应正确报告进度(self):
        """应正确调用回调函数报告进度"""
        callback = Mock()
        reporter = ThreadSafeProgressReporter(callback)

        info = ProgressInfo(
            current_row=10,
            total_rows=100,
            current_action="加载图片中"
        )

        reporter.report(info)
        callback.assert_called_once_with(info)

    def test_无回调时应安全处理(self):
        """无回调时不应抛出异常"""
        reporter = ThreadSafeProgressReporter(None)

        info = ProgressInfo(
            stage=ProgressStage.LOADING_IMAGES,
            current=10,
            total=100,
            message="加载图片中"
        )

        # 不应抛出异常
        reporter.report(info)

    def test_应限制报告频率(self):
        """应限制报告频率（节流）"""
        callback = Mock()
        reporter = ThreadSafeProgressReporter(callback)

        info = ProgressInfo(
            stage=ProgressStage.LOADING_IMAGES,
            current=10,
            total=100,
            message="加载图片中"
        )

        # 快速连续报告多次
        reporter.report(info)
        reporter.report(info)
        reporter.report(info)

        # 由于节流，应该只调用一次
        assert callback.call_count == 1

    def test_回调异常应被捕获(self):
        """回调函数抛出异常时不应影响报告器"""
        callback = Mock(side_effect=Exception("回调错误"))
        reporter = ThreadSafeProgressReporter(callback)

        info = ProgressInfo(
            stage=ProgressStage.LOADING_IMAGES,
            current=10,
            total=100,
            message="加载图片中"
        )

        # 不应抛出异常
        reporter.report(info)


class TestProcessOrchestrator:
    """流程编排器测试"""

    @pytest.fixture
    def orchestrator(self):
        """创建编排器实例"""
        return ProcessOrchestrator()

    @pytest.fixture
    def sample_excel(self, tmp_path):
        """创建示例 Excel 文件"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # 添加表头
        ws.cell(row=1, column=1, value="商品编码")
        ws.cell(row=1, column=2, value="名称")

        # 添加数据
        ws.cell(row=2, column=1, value="C001")
        ws.cell(row=2, column=2, value="产品 A")

        excel_path = tmp_path / "test.xlsx"
        wb.save(str(excel_path))

        return str(excel_path)

    @pytest.fixture
    def sample_images(self, tmp_path):
        """创建示例图片目录"""
        from PIL import Image

        image_dir = tmp_path / "images"
        image_dir.mkdir()

        # 创建测试图片
        for code in ["C001", "C002"]:
            img = Image.new('RGB', (100, 100), color='red')
            img.save(str(image_dir / f"{code}.png"))

        return str(image_dir)

    def test_应正确初始化(self, orchestrator):
        """应正确初始化编排器"""
        assert orchestrator is not None
        assert hasattr(orchestrator, 'process')

    def test_应处理基本流程(self, orchestrator, sample_excel, sample_images):
        """应处理基本的图片插入流程"""
        progress_callback = Mock()

        result = orchestrator.process(
            excel_path=sample_excel,
            image_source=sample_images,
            output_path=None,
            progress_callback=progress_callback
        )

        assert result is not None
        assert isinstance(result, ProcessResult)

    def test_应正确报告进度(self, orchestrator, sample_excel, sample_images):
        """应正确调用进度回调"""
        progress_callback = Mock()

        orchestrator.process(
            excel_path=sample_excel,
            image_source=sample_images,
            output_path=None,
            progress_callback=progress_callback
        )

        # 应该至少调用一次进度回调
        assert progress_callback.call_count > 0

    def test_应处理不存在的Excel文件(self, orchestrator):
        """应正确处理不存在的 Excel 文件"""
        with pytest.raises(Exception):
            orchestrator.process(
                excel_path="/nonexistent/path.xlsx",
                image_source="/tmp/images",
                output_path=None,
                progress_callback=None
            )

    def test_应处理不存在的图片目录(self, orchestrator, sample_excel):
        """应正确处理不存在的图片目录"""
        result = orchestrator.process(
            excel_path=sample_excel,
            image_source="/nonexistent/images",
            output_path=None,
            progress_callback=None
        )

        # 应该返回结果，但可能没有匹配的图片
        assert result is not None

    def test_应支持自定义输出路径(self, orchestrator, sample_excel, sample_images, tmp_path):
        """应支持自定义输出路径"""
        output_path = str(tmp_path / "output.xlsx")

        result = orchestrator.process(
            excel_path=sample_excel,
            image_source=sample_images,
            output_path=output_path,
            progress_callback=None
        )

        assert result is not None
        # 输出文件应该被创建
        assert Path(output_path).exists()

    def test_应处理日志回调(self, orchestrator, sample_excel, sample_images):
        """应正确调用日志回调"""
        log_callback = Mock()

        result = orchestrator.process(
            excel_path=sample_excel,
            image_source=sample_images,
            output_path=None,
            progress_callback=None,
            log_callback=log_callback
        )

        # 应该至少调用一次日志回调
        assert log_callback.call_count > 0

    def test_应处理空Excel文件(self, orchestrator, tmp_path):
        """应正确处理空 Excel 文件"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        # 不添加任何数据

        excel_path = tmp_path / "empty.xlsx"
        wb.save(str(excel_path))

        result = orchestrator.process(
            excel_path=str(excel_path),
            image_source="/tmp/images",
            output_path=None,
            progress_callback=None
        )

        assert result is not None

    def test_应正确统计处理结果(self, orchestrator, sample_excel, sample_images):
        """应正确统计处理结果"""
        result = orchestrator.process(
            excel_path=sample_excel,
            image_source=sample_images,
            output_path=None,
            progress_callback=None
        )

        # 应该有统计信息
        assert hasattr(result, 'total_rows')
        assert hasattr(result, 'matched_images')

    def test_应处理大量数据(self, orchestrator, tmp_path):
        """应正确处理大量数据"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # 添加大量数据
        ws.cell(row=1, column=1, value="商品编码")
        for i in range(2, 102):  # 100 行数据
            ws.cell(row=i, column=1, value=f"C{i:03d}")

        excel_path = tmp_path / "large.xlsx"
        wb.save(str(excel_path))

        result = orchestrator.process(
            excel_path=str(excel_path),
            image_source="/tmp/images",
            output_path=None,
            progress_callback=None
        )

        assert result is not None
        assert result.total_rows == 100


class TestProcessContext:
    """处理上下文测试"""

    def test_应正确创建上下文(self):
        """应正确创建处理上下文"""
        matcher = Mock()
        reporter = Mock()
        result = ProcessResult()

        context = ProcessContext(
            excel_path="/path/to/excel.xlsx",
            image_source="/path/to/images",
            output_path="/path/to/output.xlsx",
            matcher=matcher,
            reporter=reporter,
            log_callback=None,
            result=result
        )

        assert context.excel_path == "/path/to/excel.xlsx"
        assert context.image_source == "/path/to/images"
        assert context.output_path == "/path/to/output.xlsx"
        assert context.matcher == matcher
        assert context.reporter == reporter
        assert context.log_callback is None
        assert context.result == result
