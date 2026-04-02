"""
集成测试：测试黄色高亮功能的完整流程
"""

import sys
from pathlib import Path

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

import pytest
from openpyxl import Workbook, load_workbook
from src.core.process_engine import ProcessEngine
from src.core.image_processor import ImageProcessor


class TestHighlightIntegration:
    """黄色高亮功能集成测试"""
    
    @pytest.fixture
    def excel_with_empty_products(self, tmp_path):
        """创建包含未匹配图片商品的 Excel 文件"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # 添加表头
        ws.cell(row=1, column=1, value="序号")
        ws.cell(row=1, column=2, value="商品编码")
        ws.cell(row=1, column=3, value="名称")
        
        # 添加数据
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value="C001")  # 有图片
        ws.cell(row=2, column=3, value="产品 A")
        
        ws.cell(row=3, column=1, value=2)
        ws.cell(row=3, column=2, value="C002")  # 无图片
        ws.cell(row=3, column=3, value="产品 B")
        
        ws.cell(row=4, column=1, value=3)
        ws.cell(row=4, column=2, value="C003")  # 有图片
        ws.cell(row=4, column=3, value="产品 C")
        
        ws.cell(row=5, column=1, value=4)
        ws.cell(row=5, column=2, value="C004")  # 无图片
        ws.cell(row=5, column=3, value="产品 D")
        
        excel_path = tmp_path / "test_highlight.xlsx"
        wb.save(str(excel_path))
        
        return excel_path
    
    @pytest.fixture
    def sample_images(self, tmp_path):
        """创建示例图片"""
        from PIL import Image, ImageDraw
        
        images_dir = tmp_path / "images"
        images_dir.mkdir()
        
        # 创建 C001 的图片
        img1 = Image.new('RGB', (200, 150), color='red')
        img1.save(str(images_dir / "C001-01.jpg"))
        
        # 创建 C003 的图片
        img3 = Image.new('RGB', (200, 150), color='blue')
        img3.save(str(images_dir / "C003-01.jpg"))
        
        return images_dir
    
    def test_highlight_empty_products(self, excel_with_empty_products, sample_images, tmp_path):
        """测试完整流程中未匹配图片的商品会被高亮"""
        engine = ProcessEngine()
        
        # 执行处理
        output_path = tmp_path / "output.xlsx"
        result = engine.process(
            excel_path=str(excel_with_empty_products),
            image_source=str(sample_images),
            output_path=str(output_path)
        )
        
        # 验证处理成功
        assert result.success is True
        
        # 读取输出文件验证高亮
        wb = load_workbook(str(output_path))
        ws = wb["Sheet1"]
        
        # 商品编码在第 2 列
        product_code_column = 2
        
        # C001 (第 2 行) - 有图片，不应该高亮
        cell_row2 = ws.cell(row=2, column=product_code_column)
        # 检查是否无填充或填充为透明
        assert cell_row2.fill.start_color.rgb == "00000000"
        
        # C002 (第 3 行) - 无图片，应该黄色高亮
        cell_row3 = ws.cell(row=3, column=product_code_column)
        assert cell_row3.fill.start_color.rgb == "00FFFF00"  # 黄色
        assert cell_row3.fill.fill_type == "solid"
        
        # C003 (第 4 行) - 有图片，不应该高亮
        cell_row4 = ws.cell(row=4, column=product_code_column)
        assert cell_row4.fill.start_color.rgb == "00000000"
        
        # C004 (第 5 行) - 无图片，应该黄色高亮
        cell_row5 = ws.cell(row=5, column=product_code_column)
        assert cell_row5.fill.start_color.rgb == "00FFFF00"  # 黄色
        assert cell_row5.fill.fill_type == "solid"
