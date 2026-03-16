"""
Excel 格式设置单元测试

测试内容：
1. 单元格大小调整（列宽和行高）
2. 字体设置为微软雅黑
3. Picture 列居中对齐
"""

import pytest
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from src.core.excel_processor import ExcelProcessor


class TestExcelFormat:
    """Excel 格式设置测试类"""
    
    @pytest.fixture
    def sample_excel(self, tmp_path):
        """创建示例 Excel 文件"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # 添加表头
        ws.cell(row=1, column=1, value="序号")
        ws.cell(row=1, column=2, value="商品编码")
        ws.cell(row=1, column=3, value="名称")
        
        # 添加数据
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value="C001")
        ws.cell(row=2, column=3, value="产品 A")
        
        ws.cell(row=3, column=1, value=2)
        ws.cell(row=3, column=2, value="C002")
        ws.cell(row=3, column=3, value="产品 B")
        
        excel_path = tmp_path / "test.xlsx"
        wb.save(str(excel_path))
        
        return excel_path
    
    @pytest.fixture
    def sample_image(self, tmp_path):
        """创建示例图片"""
        from PIL import Image
        img = Image.new('RGB', (180, 138), color='red')
        image_path = tmp_path / "test_image.jpg"
        img.save(str(image_path))
        return image_path
    
    def test_add_picture_columns_font(self, sample_excel):
        """测试添加 Picture 列时设置字体为微软雅黑"""
        processor = ExcelProcessor(str(sample_excel), read_only=False)
        processor.find_sheet_with_product_code()
        
        # 添加 Picture 列
        result = processor.add_picture_columns()
        
        # 验证字体设置
        worksheet = processor.workbook["Sheet1"]
        for picture_name, column_num in result.items():
            cell = worksheet.cell(row=1, column=column_num)
            
            # 检查字体名称
            assert cell.font.name == '微软雅黑', f"{picture_name} 的字体不是微软雅黑"
            # 检查字体大小
            assert cell.font.size == 11, f"{picture_name} 的字体大小不是 11"
    
    def test_add_picture_columns_alignment(self, sample_excel):
        """测试添加 Picture 列时设置居中对齐"""
        processor = ExcelProcessor(str(sample_excel), read_only=False)
        processor.find_sheet_with_product_code()
        
        # 添加 Picture 列
        result = processor.add_picture_columns()
        
        # 验证对齐设置
        worksheet = processor.workbook["Sheet1"]
        for picture_name, column_num in result.items():
            cell = worksheet.cell(row=1, column=column_num)
            
            # 检查水平对齐
            assert cell.alignment.horizontal == 'center', f"{picture_name} 不是水平居中"
            # 检查垂直对齐
            assert cell.alignment.vertical == 'center', f"{picture_name} 不是垂直居中"
    
    def test_embed_image_cell_size(self, sample_excel, sample_image):
        """测试嵌入图片时调整单元格大小"""
        processor = ExcelProcessor(str(sample_excel), read_only=False)
        processor.find_sheet_with_product_code()
        
        # 添加 Picture 列
        result = processor.add_picture_columns()
        
        # 嵌入图片
        processor.embed_image(2, result["Picture 1"], str(sample_image))
        
        # 验证单元格大小调整
        worksheet = processor.workbook["Sheet1"]
        col_letter = worksheet.cell(row=1, column=result["Picture 1"]).column_letter
        
        # 检查列宽（180 像素 * 0.15 ≈ 27）
        column_width = worksheet.column_dimensions[col_letter].width
        assert column_width > 25, f"列宽 {column_width} 不符合预期（应该约 27）"
        
        # 检查行高（138 像素 * 0.75 ≈ 103.5）
        row_height = worksheet.row_dimensions[2].height
        assert row_height > 100, f"行高 {row_height} 不符合预期（应该约 103.5）"
    
    def test_embed_image_custom_size(self, sample_excel, sample_image):
        """测试嵌入自定义尺寸图片时调整单元格大小"""
        processor = ExcelProcessor(str(sample_excel), read_only=False)
        processor.find_sheet_with_product_code()
        
        # 添加 Picture 列
        result = processor.add_picture_columns()
        
        # 使用自定义尺寸嵌入图片
        custom_width = 200
        custom_height = 150
        processor.embed_image(
            2, 
            result["Picture 1"], 
            str(sample_image),
            width=custom_width,
            height=custom_height
        )
        
        # 验证单元格大小调整
        worksheet = processor.workbook["Sheet1"]
        col_letter = worksheet.cell(row=1, column=result["Picture 1"]).column_letter
        
        # 检查列宽（200 像素 * 0.15 = 30）
        column_width = worksheet.column_dimensions[col_letter].width
        assert column_width > 28, f"列宽 {column_width} 不符合预期（应该约 30）"
        
        # 检查行高（150 像素 * 0.75 = 112.5）
        row_height = worksheet.row_dimensions[2].height
        assert row_height > 110, f"行高 {row_height} 不符合预期（应该约 112.5）"
