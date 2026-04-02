"""
Excel 处理器模块

负责读取、处理和保存 Excel 文件
支持动态表头识别、Picture 列管理和图片嵌入

功能：
1. 读取 Excel 文件（支持 read_only 模式）
2. 动态识别表头位置（遍历查找"商品编码"所在行）
3. 处理合并单元格、不规则格式
4. 智能检查并添加 Picture 1/2/3 列
5. 图片嵌入（支持文件路径和二进制数据）
6. 保存带图 Excel 文件

重构说明：
- SheetInfo 保留在本地（仅被 ExcelProcessor 使用）
- ProgressInfo 已移至 models.py（共享数据类）
"""

import os
import platform
from pathlib import Path
from types import TracebackType
from typing import Optional, List, Union, Any
from dataclasses import dataclass, field

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.styles import Font, Alignment, PatternFill
from PIL import Image as PILImage
from io import BytesIO

from .picture_variant import VariantRecognizer, PictureColumnMapper, ColumnAdditionResult, PictureColumn
from utils.config import ImageConfig


def _get_cross_platform_font() -> str:
    """获取跨平台兼容的默认字体"""
    system = platform.system()
    if system == 'Windows':
        return '微软雅黑'
    elif system == 'Darwin':
        return 'Helvetica Neue'
    else:
        return 'Liberation Sans'


DEFAULT_FONT_NAME = _get_cross_platform_font()


# ============ 模块级辅助函数 ============

def _calculate_start_column(has_content: bool, last_column: int) -> int:
    """
    计算 Picture 列的起始列号

    Args:
        has_content: 最后一列是否有内容
        last_column: 当前最后一列

    Returns:
        int: 起始列号
    """
    if has_content:
        return last_column + 1
    else:
        return last_column


def _load_image_bytes(source: Union[str, bytes]) -> bytes:
    """
    加载图片字节数据

    Args:
        source: 图片源（文件路径或字节数据）

    Returns:
        bytes: 图片字节数据

    Raises:
        FileNotFoundError: 文件不存在
        ValueError: 无效的图片源
    """
    if isinstance(source, str):
        if not os.path.exists(source):
            raise FileNotFoundError(f"图片文件不存在：{source}")
        with open(source, 'rb') as f:
            return f.read()
    elif isinstance(source, bytes):
        return source
    else:
        raise ValueError("无效的图片源，必须是文件路径（str）或二进制数据（bytes）")


def _prepare_image_for_excel(image_bytes: bytes) -> bytes:
    """
    准备 Excel 格式的图片数据（保持 PNG 或转换为 JPEG）

    Args:
        image_bytes: 原始图片字节数据

    Returns:
        bytes: Excel 兼容的图片字节数据
    """
    pil_image = PILImage.open(BytesIO(image_bytes))
    original_format = pil_image.format or 'JPEG'

    buffer = BytesIO()
    if original_format.upper() == 'PNG':
        pil_image.save(buffer, format='PNG')
    else:
        pil_image.save(buffer, format='JPEG', quality=95)
    buffer.seek(0)
    return buffer.getvalue()


def _create_image_anchor(
        column_num: int,
        row: int,
        width: int,
        height: int
) -> TwoCellAnchor:
    """
    创建图片单元格锚点

    Args:
        column_num: 列号
        row: 行号
        width: 图片宽度
        height: 图片高度

    Returns:
        TwoCellAnchor: 单元格锚点
    """
    anchor = TwoCellAnchor()
    anchor._from = AnchorMarker(col=column_num - 1, row=row - 1, colOff=0, rowOff=0)
    anchor.to = AnchorMarker(
        col=column_num - 1,
        row=row - 1,
        colOff=pixels_to_EMU(width),
        rowOff=pixels_to_EMU(height)
    )
    anchor.editAs = 'oneCell'
    return anchor


def _add_image_to_worksheet(
        worksheet,
        xl_image: XLImage,
        column_num: int,
        row: int,
        width: int,
        height: int,
        column_width_ratio: float,
        row_height_ratio: float
) -> None:
    """
    将图片添加到工作表

    Args:
        worksheet: 工作表对象
        xl_image: Excel 图片对象
        column_num: 列号
        row: 行号
        width: 图片宽度
        height: 图片高度
        column_width_ratio: 列宽比例
        row_height_ratio: 行高比例
    """
    col_letter = get_column_letter(column_num)
    worksheet.column_dimensions[col_letter].width = width * column_width_ratio
    worksheet.row_dimensions[row].height = height * row_height_ratio
    worksheet.add_image(xl_image)


@dataclass
class SheetInfo:
    """
    工作表信息数据类

    Attributes:
        name (str): 工作表名称
        header_row (int): 表头行号（从 1 开始）
        product_code_column (int): 商品编码列号（从 1 开始）
        last_column (int): 最后一列的列号
        data_rows (List[int]): 数据行号列表
    """
    name: str
    header_row: int
    product_code_column: int
    last_column: int
    data_rows: List[int] = field(default_factory=list)


class ExcelProcessor:
    """
    Excel 处理器类

    功能：
    1. 读取 Excel 文件（read_only 模式）
    2. 动态识别表头位置
    3. 识别"商品编码"列
    4. 智能添加 Picture 1/2/3 列
    5. 嵌入图片到单元格
    6. 保存带图 Excel 文件

    Example:
        >>> processor = ExcelProcessor('/path/to/file.xlsx', read_only=True)
        >>> sheet_info = processor.find_sheet_with_product_code()
        >>> processor.add_picture_columns()
        >>> processor.embed_image(2, 'Picture 1', image_data)
        >>> processor.save('/path/to/output.xlsx')
    """

    PRODUCT_CODE_COLUMN = '商品编码'
    PICTURE_COLUMNS = ['Picture 1', 'Picture 2', 'Picture 3']
    IMAGE_WIDTH = ImageConfig.WIDTH
    IMAGE_HEIGHT = ImageConfig.HEIGHT

    # 列宽和行高调整比例（用于适配图片尺寸）
    COLUMN_WIDTH_RATIO = 0.15
    ROW_HEIGHT_RATIO = 0.75

    def __init__(self, file_path: str, read_only: bool = True):
        self.file_path = Path(file_path)
        self.read_only = read_only
        self._validate_file()
        self.workbook = load_workbook(
            filename=self.file_path,
            read_only=read_only,
            data_only=True
        )
        self.sheet_info: Optional[SheetInfo] = None
        self.output_path: Optional[Path] = None

    def _validate_file(self):
        """验证 Excel 文件"""
        if not self.file_path.exists():
            raise FileNotFoundError(f"文件不存在：{self.file_path}")

        if not self.file_path.is_file():
            raise ValueError(f"路径不是文件：{self.file_path}")

        if self.file_path.suffix.lower() not in ['.xlsx']:
            raise ValueError(
                f"不支持的文件格式：{self.file_path.suffix}，"
                f"仅支持 .xlsx 格式（Excel 2007+）"
            )

    def find_sheet_with_product_code(self) -> Optional[SheetInfo]:
        """查找包含"商品编码"列的工作表"""
        if self.sheet_info is not None:
            return self.sheet_info

        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]

            header_row = None
            product_code_column = None

            for row_idx, row in enumerate(worksheet.iter_rows(max_row=100), start=1):
                for cell in row:
                    if cell.value and self.PRODUCT_CODE_COLUMN in str(cell.value):
                        header_row = row_idx
                        product_code_column = cell.column
                        break

                if header_row:
                    break

            if header_row and product_code_column:
                last_column = worksheet.max_column

                data_rows = []
                for row_idx in range(header_row + 1, worksheet.max_row + 1):
                    row_has_data = False
                    first_col_value = None
                    for col_idx in range(1, last_column + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            if first_col_value is None:
                                first_col_value = str(cell.value).strip()
                            row_has_data = True
                            break

                    if row_has_data and first_col_value:
                        is_summary_row = any(
                            keyword in first_col_value.upper()
                            for keyword in ['TOTAL', '总计', '合计', 'SUM']
                        )
                        if not is_summary_row:
                            data_rows.append(row_idx)

                self.sheet_info = SheetInfo(
                    name=sheet_name,
                    header_row=header_row,
                    product_code_column=product_code_column,
                    last_column=last_column,
                    data_rows=data_rows
                )

                return self.sheet_info

        return None

    def _check_last_column_content(self) -> bool:
        """检查表头行最后一列在数据行中是否有内容"""
        if not self.sheet_info:
            return False

        worksheet = self.workbook[self.sheet_info.name]
        last_column = self.sheet_info.last_column

        for row_idx in self.sheet_info.data_rows:
            cell = worksheet.cell(row=row_idx, column=last_column)
            if cell.value is not None and str(cell.value).strip():
                return True

        return False

    def _column_exists(self, column_name: str) -> Optional[int]:
        """检查指定列名的列是否存在"""
        if not self.sheet_info:
            return None

        worksheet = self.workbook[self.sheet_info.name]
        recognizer = VariantRecognizer()

        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=self.sheet_info.header_row, column=col_idx)
            if cell.value:
                cell_value = str(cell.value).strip()

                if cell_value == column_name:
                    return col_idx

                result = recognizer.recognize(cell_value)
                if result:
                    base_word, number = result
                    expected_name = f"{base_word} {number}"
                    if expected_name == column_name:
                        return col_idx

        return None

    def _determine_max_pictures(
            self,
            image_matcher: Optional[Any],
            max_pictures: Optional[int]
    ) -> int:
        """
        确定最大图片数量

        Args:
            image_matcher: 图片匹配器
            max_pictures: 指定的最大图片数

        Returns:
            int: 最大图片数量
        """
        if max_pictures is not None:
            return min(max_pictures, 10)
        elif image_matcher:
            final_max = self.scan_product_images(image_matcher)
            return min(final_max, 10)
        else:
            return 3

    def _extract_base_word_from_existing_columns(self, mapper: PictureColumnMapper) -> str:
        """
        从已存在的列中提取基础列名

        Args:
            mapper: 列映射器

        Returns:
            str: 基础列名（如 "Picture"）
        """
        base_word = "Picture"
        if mapper.existing_columns:
            for number in mapper.existing_columns:
                orig = mapper.original_headers.get(number, "")
                if orig:
                    base = orig.strip()
                    for i in range(1, 20):
                        base = base.replace(str(i), "").strip()
                    if base:
                        base_word = base
                        break
        return base_word

    def _add_new_picture_columns(
            self,
            worksheet,
            needed_numbers: List[int],
            start_column: int,
            base_word: str
    ) -> List[PictureColumn]:
        """
        添加新的 Picture 列

        Args:
            worksheet: 工作表对象
            needed_numbers: 需要添加的列编号列表
            start_column: 起始列号
            base_word: 列名前缀

        Returns:
            List[PictureColumn]: 添加的列信息列表
        """
        added_columns = []

        for idx, number in enumerate(needed_numbers):
            column_num = start_column + idx
            picture_name = f"{base_word} {number}"

            cell = worksheet.cell(row=self.sheet_info.header_row, column=column_num)
            cell.value = picture_name
            cell.font = Font(name=DEFAULT_FONT_NAME, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            self.sheet_info.last_column = column_num

            added_columns.append(PictureColumn(
                number=number,
                original_header=picture_name,
                column_index=column_num,
                is_existing=False
            ))

        return added_columns

    def _rename_existing_columns(
            self,
            worksheet,
            mapper: PictureColumnMapper,
            base_word: str
    ) -> None:
        """
        重命名已存在的 Picture 列

        Args:
            worksheet: 工作表对象
            mapper: 列映射器
            base_word: 列名前缀
        """
        if not base_word:
            return

        for number, col_idx in mapper.existing_columns.items():
            old_header = mapper.original_headers.get(number, "")
            new_header = f"{base_word} {number}"
            if old_header != new_header:
                worksheet.cell(row=self.sheet_info.header_row, column=col_idx).value = new_header
                mapper.original_headers[number] = new_header

    def add_picture_columns(
            self,
            image_matcher: Optional[Any] = None,
            max_pictures: Optional[int] = None
    ) -> ColumnAdditionResult:
        """添加 Picture 列"""
        if not self.sheet_info:
            raise ValueError("未找到包含商品编码的工作表")

        worksheet = self.workbook[self.sheet_info.name]

        # 扫描现有列
        mapper = PictureColumnMapper()
        mapper.scan_worksheet(worksheet, self.sheet_info.header_row)

        # 确定基础列名
        base_word = self._extract_base_word_from_existing_columns(mapper)

        # 确定最大图片数量
        final_max = self._determine_max_pictures(image_matcher, max_pictures)
        if final_max < 1:
            final_max = 1

        # 计算需要添加的列
        needed_numbers = mapper.calculate_needed_columns(final_max)

        # 计算起始列号
        start_column = self.sheet_info.last_column
        if needed_numbers:
            has_content = self._check_last_column_content()
            start_column = _calculate_start_column(has_content, self.sheet_info.last_column)

        # 添加新列
        added_columns = self._add_new_picture_columns(
            worksheet, needed_numbers, start_column, base_word
        )

        # 重命名已存在列
        self._rename_existing_columns(worksheet, mapper, base_word)

        # 更新映射器中的列信息
        for idx, number in enumerate(needed_numbers):
            column_num = start_column + idx
            picture_name = f"{base_word} {number}"
            mapper.existing_columns[number] = column_num
            mapper.original_headers[number] = picture_name

        # 获取最终的已存在列
        final_existing_columns = mapper.to_picture_columns()

        return ColumnAdditionResult(
            added_columns=added_columns,
            existing_columns=final_existing_columns,
            total_columns=len(final_existing_columns) + len(added_columns)
        )

    def embed_image(
            self,
            row: int,
            column: Union[int, str],
            source: Union[str, bytes],
            width: int = ImageConfig.WIDTH,
            height: int = ImageConfig.HEIGHT
    ) -> None:
        """嵌入图片到指定单元格，图片随单元格调整大小（原图尺寸）"""
        if not self.sheet_info:
            raise ValueError("未找到包含商品编码的工作表")

        worksheet = self.workbook[self.sheet_info.name]

        # 解析列号
        if isinstance(column, str):
            column_num = column_index_from_string(column)
        else:
            column_num = column

        # 加载并准备图片
        image_bytes = _load_image_bytes(source)
        image_bytes = _prepare_image_for_excel(image_bytes)

        # 创建 Excel 图片对象
        xl_image = XLImage(BytesIO(image_bytes))

        # 创建锚点并设置位置
        two_cell_anchor = _create_image_anchor(column_num, row, width, height)
        xl_image.anchor = two_cell_anchor

        # 添加到工作表
        _add_image_to_worksheet(
            worksheet, xl_image, column_num, row, width, height,
            self.COLUMN_WIDTH_RATIO, self.ROW_HEIGHT_RATIO
        )

    def get_product_code(self, row: int) -> Optional[str]:
        """获取指定行的商品编码"""
        if not self.sheet_info:
            return None

        worksheet = self.workbook[self.sheet_info.name]
        cell = worksheet.cell(row=row, column=self.sheet_info.product_code_column)

        if cell.value:
            return str(cell.value).strip()

        return None

    def get_product_codes(self) -> List[str]:
        """获取所有商品编码"""
        if not self.sheet_info:
            return []

        codes = []

        for row in self.sheet_info.data_rows:
            code = self.get_product_code(row)
            if code:
                codes.append(code)

        return codes

    def scan_product_images(self, image_matcher: Any) -> int:
        """扫描商品的图片数量"""
        if not self.sheet_info:
            return 0

        max_pictures = 0

        for row in self.sheet_info.data_rows:
            product_code = self.get_product_code(row)
            if not product_code:
                continue

            picture_count = 0
            for pic_num in range(1, 11):
                if image_matcher.has_image(product_code, pic_num):
                    picture_count += 1
                else:
                    break

            max_pictures = max(max_pictures, picture_count)

        return max_pictures

    def save(self, output_path: Optional[str] = None) -> Path:
        """保存 Excel 文件"""
        if output_path:
            self.output_path = Path(output_path)
        else:
            stem = self.file_path.stem
            suffix = self.file_path.suffix

            target_suffix = "_含图"
            while stem.endswith(target_suffix):
                stem = stem[:-len(target_suffix)]

            self.output_path = self.file_path.parent / f"{stem}{target_suffix}{suffix}"

        self.workbook.save(str(self.output_path))

        return self.output_path

    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None

    def __enter__(self) -> 'ExcelProcessor':
        """上下文管理器入口"""
        return self

    def __exit__(
            self,
            exc_type: Optional[type],
            exc_val: Optional[BaseException],
            exc_tb: Optional[TracebackType]
    ) -> None:
        """上下文管理器出口"""
        self.close()

    def highlight_empty_product_codes(
            self,
            rows: List[int],
            fill_color: str = "FFFF00"
    ) -> None:
        """高亮显示未匹配到图片的商品编码单元格"""
        if not self.sheet_info:
            raise ValueError("未找到包含商品编码的工作表")

        worksheet = self.workbook[self.sheet_info.name]

        yellow_fill = PatternFill(
            start_color=fill_color,
            end_color=fill_color,
            fill_type="solid"
        )

        for row in rows:
            cell = worksheet.cell(
                row=row,
                column=self.sheet_info.product_code_column
            )
            cell.fill = yellow_fill

            if cell.font is None or cell.font.color is None:
                cell.font = Font(color="000000")
